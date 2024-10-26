import os
import openpyxl
import numpy as np

# 参数设置
i_max = 54 + 1  # 土地编号
j_max = 41 + 1  # 作物编号
n_max = 2 + 1  # 季度
k_max = 8 + 1  # 年数


# 定义土地类
class Land:
    def __init__(self, name, no, size, status, mark, j_max, n_max):
        self.name = name  # 土地名称
        self.no = no  # 土地编号
        self.size = size  # 土地面积
        self.status = status  # 土地状态
        self.production = np.zeros((j_max, n_max))  # 产量二维数组
        self.cost = np.zeros((j_max, n_max))  # 成本二维数组
        self.sale = np.zeros((j_max, n_max))  # 销售二维数组
        self.mark = mark  # 土地的唯一标号 如B2 C2

    def printself(self):
        print(f"name: {self.name},mark: {self.mark}, no: {self.no}, size: {self.size},status:{self.status}")
        print(f"production: {self.production}")
        print(f"cost: {self.cost}")
        print(f"sale: {self.sale}")


# 初始化空的列表来存储所有的 Land 实例
lands = []

# 检查路径和 Excel 文件
path = r"./data"
if not os.path.exists(path):
    print("Incorrect path!")
    exit(1)

os.chdir(path)
try:
    workbook = openpyxl.load_workbook('附件1_现有作物与土地情况.xlsx')
except FileNotFoundError:
    print("No input file found!")
    exit(1)

sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[0]]  # 读取第一个工作表
cnt = 0
# 读取 Excel 文件的内容并解析，不跳过第一行的表头 所以 land[0]弃用
for row in sheet.iter_rows(min_row=1, values_only=True):
    land_no = cnt  # 地块编号用计数器记录
    cnt += 1
    mark = row[0]
    land_status = row[1]  # 地块类型
    land_size = row[2]  # 地块面积

    # 创建一个 Land 实例
    land_instance = Land(f"Land {land_no}", land_no, land_size, land_status, mark, j_max, n_max)

    # 将该实例添加到 lands 数组中
    lands.append(land_instance)
#
# # 检查 Land 数组是否正确生成
# for land in lands:
#     print(f"Land {land.no}: Size = {land.size}, Status = {land.status} mark = {land.mark}")
# 定义作物类
sheet = workbook[sheet_names[1]]


class Corps:
    def __init__(self, name, no, status, expected_sell):
        self.name = name  # 作物名称
        self.no = no  # 作物编号
        self.expected_sell = expected_sell  # 预期销量
        self.status = status

    def printself(self):
        print(f"名字：{self.name}, {self.no} 号作物, 种类：{self.status}, 预计能售出： {self.expected_sell} 斤")


# 初始化空的列表来存储所有的 Corps 实例
corps_list = []

# 假设你已经打开了 Excel 文件并定位到相应的工作表
for row in sheet.iter_rows(min_row=1, values_only=True):  # 使用 enumerate 作为计数器
    corp_name = row[1]  # 作物名称
    corp_no = row[0]  # 预期销量
    corp_status = row[2]
    # 创建一个 Corps 实例
    corp_instance = Corps(corp_name, corp_no, corp_status, 0)

    # 将该实例添加到 corps_list 数组中
    corps_list.append(corp_instance)

# 检查作物数组是否正确生成
# for corp in corps_list:
#     print(f"Corp No: {corp.no}, Name: {corp.name}, status: {corp.status}")
#-------------------------------
#打开附件二
workbook = openpyxl.load_workbook('附件2_去年作物与收成情况.xlsx')
if (workbook == None):
    print("no input file found!")
    exit(1)

sheet_names = workbook.sheetnames  #选择一个表格名
sheet = workbook[sheet_names[1]]  #2023年统计的相关数据

for row in sheet.iter_rows(min_row=2, max_row=108, values_only=True):  #对表格的每一行读取
    # 需要再调整读入的形式
    # print(row)
    for land in lands:

        num1, num2 = row[7].split('-')
        if (land.status == row[3]):
            if (row[3] == land.status):
                if (row[4] == '单季'):
                    # print("de")
                    land.sale[row[1]][1] = (float(num1) + float(num2)) / 2.0
                    land.cost[row[1]][1] = row[6]
                    land.production[row[1]][1] = row[5]
                if (row[4] == '第一季'):
                    land.sale[row[1]][1] = (float(num1) + float(num2)) / 2.0
                    land.cost[row[1]][1] = row[6]
                    land.production[row[1]][1] = row[5]
                if (row[4] == '第二季'):
                    land.sale[row[1]][2] = (float(num1) + float(num2)) / 2.0
                    land.cost[row[1]][2] = row[6]
                    land.production[row[1]][2] = row[5]
            else:
                continue
land2 = lands[37]
for land1 in lands:

    if (land1.status == '智慧大棚'):
        for i in range(j_max):
            land1.production[i][1] = land2.production[i][1]
            land1.cost[i][1] = land2.cost[i][1]
            land1.sale[i][1] = land2.sale[i][1]
# print(lands[51].cost[27][1])
sheet = workbook[sheet_names[0]]  #2023年的种植情况
for row in sheet.iter_rows(min_row=2, max_row=88, values_only=True):
    if (row[0] != None):
        saver = row[0]
    for land in lands:
        if (land.mark == saver):
            if (row[5] == '单季'):
                corps_list[row[1]].expected_sell += row[4] * land.production[row[1]][1]
                print(f"name {land.status}+{corps_list[row[1]].name}+{row[4]}*{land.production[row[1]][1]}")
            if (row[5] == '第一季'):
                corps_list[row[1]].expected_sell += row[4] * land.production[row[1]][1]
                print(f"name {land.status}+{corps_list[row[1]].name}+{row[4]}*{land.production[row[1]][1]}")

            if (row[5] == '第二季'):
                corps_list[row[1]].expected_sell += row[4] * land.production[row[1]][2]
                print(f"name {land.status}+{corps_list[row[1]].name}+{row[4]}*{land.production[row[1]][2]}")

#如何调用？
#我们有corps_list[42](0处弃用)：作物表 包含 .name .no .status .expected_sell .printself()
#我们有lands[55](0处弃用）：土地表 包含 .name .no .status .size .mark(是土地标号) .production[][] .sale[][] .cost[][] .printself()
corps_list[16].printself()
