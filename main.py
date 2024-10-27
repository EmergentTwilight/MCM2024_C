import os
import openpyxl
import numpy as np

LAND_NUM = 54
CROP_NUM = 41
SEASON_NUM = 2
YEAR_NUM = 8  # 2023-2030

class Land:
    def __init__(self, name: str, type: str, size: float):
        self.name = name  # e.g. A1
        self.type = type  # e.g. 平旱地
        self.size = size  # e.g. 80（亩）
        self.production = np.zeros((CROP_NUM, SEASON_NUM))
        self.cost = np.zeros((CROP_NUM, SEASON_NUM))
        self.price = np.zeros((CROP_NUM, SEASON_NUM))

class Crop:
    def __init__(self, name: str, type: str):
        self.name = name  # e.g. 水稻
        self.type = type  # e.g. 粮食（豆类）
        self.expected_sale = 0  # 预期销量（斤）


crop_id_of = {}
land_id_of = {}
lands = []
crops = []

path = os.path.join(os.getcwd(), 'problem')
print(path)
if not os.path.exists(path):
    print("Incorrect path!")
    exit(1)

try:
    workbook = openpyxl.load_workbook(os.path.join(path, "附件1_现有作物与土地情况.xlsx"))
except FileNotFoundError:
    print("Attachment 1 not found!")
    exit(1)

sheet_names = workbook.sheetnames

# 读取第一张工作表
sheet = workbook[sheet_names[0]]
cnt = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    land_name = row[0].strip()
    land_type = row[1].strip()
    land_size = row[2]
    land_id_of[land_name] = cnt

    land_instance = Land(land_name, land_type, land_size)
    lands.append(land_instance)
    cnt += 1

# for land in lands:
#     print(f'Land name: {land.name}, type: {land.type}, size: {land.size}, id: {land_id_of[land.name]}')

sheet = workbook[sheet_names[1]]
cnt = 0
for row in sheet.iter_rows(min_row=2, max_row=42, values_only=True):
    crop_name = row[1].strip()
    crop_id_of[crop_name] = cnt
    crop_type = row[2].strip()
    crop_instance = Crop(crop_name, crop_type)
    crops.append(crop_instance)
    cnt += 1

# for crop in crops:
#     print(f'Crop name: {crop.name}, type: {crop.type}, id: {crop_id_of[crop.name]}')
workbook.close()

try:
    workbook = openpyxl.load_workbook(os.path.join(path, "附件2_去年作物与收成情况.xlsx"))
except FileNotFoundError:
    print("Attachment 2 not found!")
    exit(1)

sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[1]]  # 作物产量、单价等数据

for row in sheet.iter_rows(min_row=2, max_row=108, values_only=True):
    crop_name = row[2].strip()  # 作物名称
    land_type = row[3].strip()  # 地块类型
    crop_season = row[4]  # 种植季次
    crop_production = row[5]  # 亩产量 斤/亩
    crop_cost = row[6]  # 种植成本 元/亩
    price1, price2 = row[7].split('-')  # 销售单价 元/斤

    price = (float(price1) + float(price2)) / 2  # 价格暂时取平均数

    for land in lands:
        if land.type == land_type:
            if crop_season == '单季' or crop_season == '第一季':
                land.production[crop_id_of[crop_name], 0] = crop_production
                land.cost[crop_id_of[crop_name], 0] = crop_cost
                land.price[crop_id_of[crop_name], 0] = price
            elif crop_season == '第二季':
                land.production[crop_id_of[crop_name], 0] = crop_production
                land.cost[crop_id_of[crop_name], 0] = crop_cost
                land.price[crop_id_of[crop_name], 0] = price

# 表格中缺了智慧大棚第一季的数据，需要手动填写
format_land = lands[land_id_of['E1']]  # 找普通大棚当作模板
for land_id in range(land_id_of['F1'], land_id_of['F4'] + 1):  # 智慧大棚
    land = lands[land_id]
    for crop_id in range(CROP_NUM):
        land.production[crop_id][0] = format_land.production[crop_id][0]
        land.cost[crop_id][0] = format_land.cost[crop_id][0]
        land.price[crop_id][0] = format_land.price[crop_id][0]

sheet = workbook[sheet_names[0]]  # 2023 年种植情况
land_name_buffer = None
for row in sheet.iter_rows(min_row=2, max_row=88, values_only=True):
    if row[0] is not None:
        land_name_buffer = row[0]
    crop_name = row[2].strip()
    crop_area = row[4]
    this_crop = crops[crop_id_of[crop_name]]
    this_land = lands[land_id_of[land_name_buffer]]
    season = row[5]
    if season == "单季" or season == "第一季":
        this_crop.expected_sale += crop_area * this_land.production[crop_id_of[crop_name], 0]
    elif season == "第二季":
        this_crop.expected_sale += crop_area * this_land.production[crop_id_of[crop_name], 1]

workbook.close()
# print(crops[crop_id_of["小麦"]].expected_sale)
# print(crops[crop_id_of["空心菜"]].expected_sale)


from gurobipy import GRB, Model

problem_type = "滞销"
# problem_type = "降价"

SubProblem1 = Model("SubProblem1")

# a_26_15_8 = SubProblem1.addVars(26, 15, 8, vtype=GRB.CONTINUOUS, name='a')  # 地，物，年
a_26_15_8 = SubProblem1.addVars(26, 15, 8, vtype=GRB.BINARY, name='a')  # 地，物，年
max_i, max_j, max_k = 26, 15, 8
bean_id_range = (1, 5)  # 0-4 是豆类植物

# # 0.1. 24-30年决策变量上下限，若使用连续变量需要确定上下界
# for i in range(26):
#     for j in range(15):
#         for k in range(1, 8):
#             a_26_15_8[i, j, k].set(GRB.Attr.LB, 0.0)
#             a_26_15_8[i, j, k].set(GRB.Attr.UB, 1.0)

# 0.2. 读入23年的数据
try:
    workbook = openpyxl.load_workbook(os.path.join(path, "附件2_去年作物与收成情况.xlsx"))
except FileNotFoundError:
    print("Attachment 2 not found!")
    exit(1)

sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[0]]  # 23年种植情况

for row in sheet.iter_rows(min_row=2, max_row=27, values_only=True):
    land_name = row[0].strip()
    crop_name = row[2].strip()
    land_id = land_id_of[land_name]
    crop_id = crop_id_of[crop_name]
    for j in range(max_j):
        SubProblem1.addConstr(a_26_15_8[land_id, j, 0] == (1 if j == crop_id else 0))

workbook.close()

# 1. 每块地刚好用满
constraint_1 = SubProblem1.addConstrs(
    sum(
        a_26_15_8[i, j, k]
        for j in range(max_j)  # 所有作物比例之和
    ) == 1
    for i in range(max_i)  # 对于每块地
    for k in range(1, max_k)  # 对于 24-30 年
)

# 2. 豆类作物三年至少种一次
constraint_2 = SubProblem1.addConstrs(
    sum(
        a_26_15_8[i, j, k] + a_26_15_8[i, j, k + 1] + a_26_15_8[i, j, k + 2]  # 连续三年之和
        for j in range(*bean_id_range)  # 豆类植物
    ) >= 1
    for i in range(max_i)  # 对于每块地
    for k in range(max_k - 2)  # 对于 23-28 年
)

# 3. 不重茬
constraint_3 = SubProblem1.addConstrs(
    a_26_15_8[i, j, k] + a_26_15_8[i, j, k + 1] <= 1
    for i in range(max_i)  # 对于每块地
    for j in range(max_j)  # 对于每种作物
    for k in range(max_k - 1)  # 对于 23-29 年
)

# 4. 目标函数
profit_expr = 0
for k in range(1, max_k):
    year = k  # 对于每一年
    for j in range(max_j):
        crop_id = j
        crop = crops[crop_id]  # 对于每种作物
        cost = 0
        production = 0
        for i in range(max_i):  # 对于每块地
            land_id = i
            land = lands[land_id]
            planted_area = a_26_15_8[i, j, k] * land.size
            production += planted_area * land.production[crop_id, 0]
            cost += planted_area * land.cost[crop_id, 0]
        sale_var = SubProblem1.addVar(vtype=GRB.CONTINUOUS, name=f'sale_{j}_{k}')

        SubProblem1.addConstr(sale_var <= production)
        SubProblem1.addConstr(sale_var <= crop.expected_sale)

        income = sale_var * lands[0].price[crop_id, 0]
        if problem_type == "降价":
            discount_sale_var = SubProblem1.addVar(vtype=GRB.CONTINUOUS, name=f'discount_sale_{j}_{k}')
            is_positive = SubProblem1.addVar(vtype=GRB.BINARY, name=f'is_positive{j}_{k}')

            # 约束 is_positive
            M = 1e10
            SubProblem1.addConstr(production - crop.expected_sale <= M * is_positive)
            SubProblem1.addConstr(production - crop.expected_sale >= -M * (1 - is_positive))

            # 约束 discount_sale_var = min(0, production - crop.expected_sale)
            SubProblem1.addConstr(discount_sale_var <= production - crop.expected_sale + M * (1 - is_positive))
            SubProblem1.addConstr(discount_sale_var <= M * is_positive)
            SubProblem1.addConstr(discount_sale_var >= production - crop.expected_sale - M * (1 - is_positive))
            SubProblem1.addConstr(discount_sale_var >= 0)

            income += discount_sale_var * lands[0].price[crop_id, 0] * 0.5  # 50% 降价出售

        profit_expr += (income - cost)

SubProblem1.setObjective(profit_expr, GRB.MAXIMIZE)

# 开始优化
SubProblem1.optimize()

# 结果输出和保存
result = np.zeros((max_i, max_j, max_k))

for i in range(max_i):
    for j in range(max_j):
        for k in range(max_k):
            print(f'a[{i}, {j}, {k}] = {a_26_15_8[i, j, k].X}')
            result[i, j, k] = a_26_15_8[i, j, k].X

import pickle as pkl

with open(f'results/SubProblem1_{problem_type}.pkl', 'wb') as f:
    pkl.dump(result, f)